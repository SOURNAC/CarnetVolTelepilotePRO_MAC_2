from __future__ import annotations

import json
import os
import shutil
import sys
import webbrowser
from dataclasses import asdict, dataclass
from datetime import date, datetime
from pathlib import Path

if getattr(sys, "frozen", False):
    base = Path(getattr(sys, "_MEIPASS", Path(sys.executable).resolve().parent))
    os.environ.setdefault("TCL_LIBRARY", str(base / "tcl" / "tcl8.6"))
    os.environ.setdefault("TK_LIBRARY", str(base / "tcl" / "tk8.6"))

from tkinter import BOTH, END, HORIZONTAL, LEFT, RIGHT, VERTICAL, X, Y, filedialog, messagebox
import tkinter as tk
from tkinter import ttk

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


APP_NAME = "Carnet de vol télépilote PRO"
YEAR = 2026
LINKS = [
    ("Alpha Tango", "https://alphatango.aviation-civile.gouv.fr/login.jsp", "#7b1fa2"),
    ("Cartes.gouv", "https://cartes.gouv.fr/decouvrir/", "#2e7d32"),
    ("SIA", "https://www.sia.aviation-civile.gouv.fr/", "#1565c0"),
    ("METAR/TAF", "https://metar-taf.com/fr?__cf_chl_tk=O0gStzi889nKzhWTXLCQ1xlwkaWd0GvROEOWEIeIdhU-1778417919-1.0.1.1-1jExqFezPXigumLp7P4HUgWdaC6dOAo14Q028YFbO5o", "#00838f"),
    ("Windy", "https://www.windy.com/?44.099,4.755,5,p:favs", "#ef6c00"),
    ("Fly by", "https://airspacedrone.com/flyby/", "#c62828"),
]
CONTACT_TEXT = "Administrateur : Guy Sournac - guy.sournac@hotmail.fr"
CONTACT_MAIL = "guy.sournac@hotmail.fr"
LETTER_TEMPLATE = "assets/lettre_mission_vierge.docx"
CERFA_TEMPLATE = "assets/cerfa_15476-04.pdf"
MISSION_GUIDE_TEMPLATE = "assets/guide_preparation_mission_complet.pdf"
CONSENT_TEMPLATE = "assets/attestation_consentement_vierge.pdf"
LOGO_IMAGE = "assets/tof_logo_header.png"
COLOR_GOLD = "#c6a35a"
COLOR_GOLD_LIGHT = "#efe3c4"
COLOR_CHARCOAL = "#3e3d3d"
COLOR_CREAM = "#faf7ee"
COLOR_BLUE_DARK = "#1f3e4e"
SCENARIOS = ["CAPTIF", "STS 02", "STS 01", "ENTRAINEMENT", "A1", "A2", "A3"]
SCENARIO_COLORS = {
    "CAPTIF": "#fff45a",
    "STS 02": "#ffc229",
    "STS 01": "#21aee4",
    "ENTRAINEMENT": "#d9d9d9",
    "A1": "#7435a6",
    "A2": "#111111",
    "A3": "#d40000",
}
DOMAINS = [
    "Photo",
    "Thermographie",
    "Inspection technique",
    "Photogrammétrie",
    "Formation",
    "Immobilier",
    "Agriculture",
    "Sûreté",
    "Évaluation",
    "Initiation",
    "Entrainement",
    "Situation normale",
    "Situation anormale",
    "Situation avec observateur",
    "Autre",
]


def app_dir() -> Path:
    if getattr(sys, "frozen", False):
        return Path(sys.executable).resolve().parent
    return Path(__file__).resolve().parent


def resource_path(relative_path: str) -> Path:
    if getattr(sys, "frozen", False):
        return Path(getattr(sys, "_MEIPASS", app_dir())) / relative_path
    return Path(__file__).resolve().parent / relative_path


def data_dir() -> Path:
    base = os.environ.get("LOCALAPPDATA")
    if base:
        return Path(base) / "CarnetVolTelepilotePRO"
    return Path.home() / "AppData" / "Local" / "CarnetVolTelepilotePRO"


def data_file() -> Path:
    return data_dir() / "carnet_vol_donnees.json"


DATA_FILE = data_file()
LEGACY_DATA_FILE = app_dir() / "carnet_vol_donnees.json"


@dataclass
class Flight:
    date: str
    aircraft: str
    scenario: str
    domain: str
    hours: float
    location: str
    notes: str
    mission: str = ""


def parse_hours(value: str) -> float:
    text = value.strip().replace(",", ".")
    if not text:
        return 0.0
    if ":" in text:
        parts = text.split(":")
        h = float(parts[0] or 0)
        m = float(parts[1] or 0) if len(parts) > 1 else 0
        return h + m / 60
    return float(text)


def fmt_hours(hours: float) -> str:
    total_minutes = int(round(hours * 60))
    h, m = divmod(total_minutes, 60)
    return f"{h}h{m:02d}"


def safe_float(value) -> float:
    try:
        return float(value)
    except Exception:
        return 0.0


def normalize_scenario(value: str) -> str:
    return "ENTRAINEMENT" if value == "INITIATION" else value


class CarnetApp(tk.Tk):
    def __init__(self) -> None:
        super().__init__()
        self.title(APP_NAME)
        self.geometry("1400x760")
        self.minsize(1180, 660)
        self.configure(bg=COLOR_CREAM)
        self.flights: list[Flight] = []
        self.edit_index: int | None = None
        self.owner_name = tk.StringVar()
        self.logo_image = None

        self._style()
        self._build()
        self.load_data()
        self.refresh()
        self.protocol("WM_DELETE_WINDOW", self.on_close)

    def _style(self) -> None:
        style = ttk.Style(self)
        try:
            style.theme_use("clam")
        except tk.TclError:
            pass
        style.configure("TFrame", background=COLOR_CREAM)
        style.configure("Card.TFrame", background="white", relief="flat")
        style.configure("Header.TFrame", background=COLOR_GOLD)
        style.configure("Blue.TFrame", background=COLOR_GOLD_LIGHT)
        style.configure("TLabel", background=COLOR_CREAM, font=("Arial", 10), foreground=COLOR_CHARCOAL)
        style.configure("Card.TLabel", background="white", font=("Arial", 10), foreground=COLOR_CHARCOAL)
        style.configure("Header.TLabel", background=COLOR_GOLD, font=("Arial", 10), foreground=COLOR_CHARCOAL)
        style.configure("Title.TLabel", background=COLOR_GOLD, font=("Arial", 18, "bold"), foreground=COLOR_CHARCOAL)
        style.configure("Footer.TLabel", background=COLOR_CREAM, font=("Arial", 8), foreground=COLOR_CHARCOAL)
        style.configure("Mail.TLabel", background=COLOR_CREAM, font=("Arial", 8, "underline"), foreground=COLOR_BLUE_DARK)
        style.configure("Kpi.TLabel", background="white", font=("Arial", 16, "bold"), foreground=COLOR_BLUE_DARK)
        style.configure("TButton", font=("Arial", 10, "bold"))
        style.configure("Treeview", font=("Arial", 9), rowheight=24)
        style.configure("Treeview.Heading", font=("Arial", 9, "bold"), background=COLOR_CHARCOAL, foreground="white")

    def _build(self) -> None:
        header = ttk.Frame(self, padding=(14, 8), style="Header.TFrame")
        header.pack(fill=X)
        try:
            self.logo_image = tk.PhotoImage(file=str(resource_path(LOGO_IMAGE)))
            self.iconphoto(False, self.logo_image)
            tk.Label(header, image=self.logo_image, bg=COLOR_GOLD, bd=0).pack(side=LEFT, padx=(0, 12))
        except Exception:
            self.logo_image = None
        ttk.Label(header, text=f"{APP_NAME} de :", style="Title.TLabel").pack(side=LEFT)
        ttk.Entry(header, textvariable=self.owner_name, width=18, font=("Arial", 13)).pack(side=LEFT, padx=(10, 10), ipady=2)

        nav = ttk.Frame(self, padding=(16, 0, 16, 8), style="Header.TFrame")
        nav.pack(fill=X)
        links = ttk.Frame(nav, style="Header.TFrame")
        links.pack(side=LEFT)
        for label, url, color in LINKS:
            tk.Button(
                links,
                text=label,
                command=lambda link=url: self.open_link(link),
                bg=color,
                fg="white",
                activebackground=color,
                activeforeground="white",
                relief="flat",
                font=("Arial", 8, "bold"),
                padx=6,
                pady=3,
            ).pack(side=LEFT, padx=2)
        actions = ttk.Frame(nav, style="Header.TFrame")
        actions.pack(side=RIGHT)
        ttk.Button(actions, text="Guide utilisation", command=self.show_user_guide).pack(side=RIGHT, padx=4)
        ttk.Button(actions, text="Attestation consentement", command=self.export_consent_template).pack(side=RIGHT, padx=4)
        ttk.Button(actions, text="Guide mission", command=self.export_mission_guide).pack(side=RIGHT, padx=4)
        ttk.Button(actions, text="Lettre mission", command=self.export_letter_template).pack(side=RIGHT, padx=4)
        ttk.Button(actions, text="CERFA", command=self.export_cerfa_template).pack(side=RIGHT, padx=4)
        ttk.Button(actions, text="Exporter Excel", command=self.export_excel).pack(side=RIGHT, padx=4)

        body = ttk.Frame(self, padding=(16, 0, 16, 12))
        body.pack(fill=BOTH, expand=True)

        form = ttk.Frame(body, style="Card.TFrame", padding=12)
        form.pack(fill=X, pady=(0, 10))
        for i in range(7):
            form.columnconfigure(i, weight=1)

        self.vars = {
            "date": tk.StringVar(value=date.today().strftime("%d/%m/%Y")),
            "aircraft": tk.StringVar(),
            "scenario": tk.StringVar(value="STS 01"),
            "domain": tk.StringVar(value="Photo"),
            "hours": tk.StringVar(value="1"),
            "location": tk.StringVar(),
            "notes": tk.StringVar(),
        }

        fields = [
            ("Date", "date", "entry"),
            ("Aéronef", "aircraft", "entry"),
            ("Scénario", "scenario", "scenario"),
            ("Domaine", "domain", "domain"),
            ("Durée (h)", "hours", "entry"),
            ("Lieu", "location", "entry"),
            ("Observations", "notes", "entry"),
        ]
        for col, (label, key, kind) in enumerate(fields):
            ttk.Label(form, text=label, style="Card.TLabel").grid(row=0, column=col, sticky="w", padx=3)
            if kind == "scenario":
                widget = ttk.Combobox(form, textvariable=self.vars[key], values=SCENARIOS, state="readonly", width=12)
            elif kind == "domain":
                widget = ttk.Combobox(form, textvariable=self.vars[key], values=DOMAINS, width=18)
            else:
                widget = ttk.Entry(form, textvariable=self.vars[key], width=16)
            widget.grid(row=1, column=col, sticky="ew", padx=3, pady=(2, 0))

        buttons = ttk.Frame(form, style="Card.TFrame")
        buttons.grid(row=2, column=0, columnspan=7, sticky="e", pady=(10, 0))
        ttk.Button(buttons, text="Ajouter / Modifier", command=self.add_or_update).pack(side=LEFT, padx=3)
        ttk.Button(buttons, text="Nouveau", command=self.clear_form).pack(side=LEFT, padx=3)
        ttk.Button(buttons, text="Supprimer", command=self.delete_selected).pack(side=LEFT, padx=3)
        ttk.Button(buttons, text="Reset", command=self.reset_all).pack(side=LEFT, padx=3)

        kpis = ttk.Frame(body)
        kpis.pack(fill=X, pady=(0, 10))
        self.kpi_total = self._kpi(kpis, "Total annuel", "0h00")
        self.kpi_count = self._kpi(kpis, "Nombre de vols", "0")

        lower = ttk.Frame(body)
        lower.pack(fill=BOTH, expand=True)

        table_frame = ttk.Frame(lower, style="Card.TFrame", padding=8)
        table_frame.pack(side=LEFT, fill=BOTH, expand=True, padx=(0, 10))

        columns = ("date", "aircraft", "scenario", "domain", "hours", "location", "notes")
        self.tree = ttk.Treeview(table_frame, columns=columns, show="headings", selectmode="browse")
        headings = {
            "date": "Date",
            "aircraft": "Aéronef",
            "scenario": "Scénario",
            "domain": "Domaine",
            "hours": "Durée",
            "location": "Lieu",
            "notes": "Observations",
        }
        widths = {"date": 90, "aircraft": 130, "scenario": 105, "domain": 190, "hours": 80, "location": 150, "notes": 260}
        for col in columns:
            self.tree.heading(col, text=headings[col])
            self.tree.column(col, width=widths[col], minwidth=50, anchor="center" if col in {"date", "scenario", "hours"} else "w")
        for scenario, color in SCENARIO_COLORS.items():
            self.tree.tag_configure(scenario, background=color, foreground="white" if scenario in {"A1", "A2", "A3"} else "black")
        yscroll = ttk.Scrollbar(table_frame, orient=VERTICAL, command=self.tree.yview)
        xscroll = ttk.Scrollbar(table_frame, orient=HORIZONTAL, command=self.tree.xview)
        self.tree.configure(yscrollcommand=yscroll.set, xscrollcommand=xscroll.set)
        self.tree.pack(side=LEFT, fill=BOTH, expand=True)
        yscroll.pack(side=RIGHT, fill=Y)
        xscroll.pack(side=tk.BOTTOM, fill=X)
        self.tree.bind("<<TreeviewSelect>>", self.load_selected)

        summary_frame = ttk.Frame(lower, style="Card.TFrame", padding=8)
        summary_frame.pack(side=RIGHT, fill=Y)
        ttk.Label(summary_frame, text="Synthèse", style="Card.TLabel", font=("Arial", 12, "bold")).pack(anchor="w", pady=(0, 6))
        ttk.Label(summary_frame, text="Par scénario", style="Card.TLabel", font=("Arial", 10, "bold")).pack(anchor="w")
        self.summary_scenario = ttk.Treeview(summary_frame, columns=("name", "hours"), show="headings", height=8)
        self.summary_scenario.heading("name", text="Scénario")
        self.summary_scenario.heading("hours", text="Heures")
        self.summary_scenario.column("name", width=140)
        self.summary_scenario.column("hours", width=70, anchor="center")
        self.summary_scenario.pack(fill=X, pady=(2, 10))
        ttk.Label(summary_frame, text="Par domaine", style="Card.TLabel", font=("Arial", 10, "bold")).pack(anchor="w")
        self.summary_domain = ttk.Treeview(summary_frame, columns=("name", "hours"), show="headings", height=10)
        self.summary_domain.heading("name", text="Domaine")
        self.summary_domain.heading("hours", text="Heures")
        self.summary_domain.column("name", width=190)
        self.summary_domain.column("hours", width=70, anchor="center")
        self.summary_domain.pack(fill=BOTH, expand=True, pady=(2, 0))

        footer = ttk.Frame(self, padding=(16, 0, 16, 6))
        footer.pack(fill=X)
        ttk.Label(footer, text="Administrateur : Guy Sournac - ", style="Footer.TLabel").pack(side=LEFT)
        mail = ttk.Label(footer, text=CONTACT_MAIL, style="Mail.TLabel", cursor="hand2")
        mail.pack(side=LEFT)
        mail.bind("<Button-1>", lambda _event: self.open_mail())

    def _kpi(self, parent, title: str, value: str) -> tk.StringVar:
        frame = ttk.Frame(parent, style="Card.TFrame", padding=12)
        frame.pack(side=LEFT, fill=X, expand=True, padx=4)
        ttk.Label(frame, text=title, style="Card.TLabel").pack(anchor="w")
        var = tk.StringVar(value=value)
        ttk.Label(frame, textvariable=var, style="Kpi.TLabel").pack(anchor="w")
        return var

    def load_data(self) -> None:
        self.migrate_legacy_data()
        if not DATA_FILE.exists():
            return
        try:
            raw = json.loads(DATA_FILE.read_text(encoding="utf-8"))
            if isinstance(raw, dict):
                self.owner_name.set(raw.get("owner_name", ""))
                rows = raw.get("flights", [])
            else:
                rows = raw
            self.flights = []
            for item in rows:
                cleaned = {
                    "date": item.get("date", ""),
                    "aircraft": item.get("aircraft", ""),
                    "scenario": normalize_scenario(item.get("scenario", "STS 01")),
                    "domain": item.get("domain", "Autre"),
                    "hours": safe_float(item.get("hours", 0)),
                    "location": item.get("location", ""),
                    "notes": item.get("notes", ""),
                    "mission": item.get("mission", ""),
                }
                self.flights.append(Flight(**cleaned))
        except Exception as exc:
            messagebox.showwarning(APP_NAME, f"Impossible de charger les données : {exc}")

    def migrate_legacy_data(self) -> None:
        if DATA_FILE.exists() or not LEGACY_DATA_FILE.exists():
            return
        try:
            DATA_FILE.parent.mkdir(parents=True, exist_ok=True)
            shutil.copyfile(LEGACY_DATA_FILE, DATA_FILE)
        except Exception:
            pass

    def _write_data(self) -> None:
        payload = {
            "owner_name": self.owner_name.get().strip(),
            "flights": [asdict(f) for f in self.flights],
        }
        DATA_FILE.parent.mkdir(parents=True, exist_ok=True)
        DATA_FILE.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")

    def save_data(self) -> None:
        self._write_data()
        messagebox.showinfo(APP_NAME, "Données enregistrées.")

    def open_link(self, url: str) -> None:
        webbrowser.open(url)

    def open_mail(self) -> None:
        webbrowser.open(f"mailto:{CONTACT_MAIL}")

    def show_user_guide(self) -> None:
        guide = tk.Toplevel(self)
        guide.title("Guide utilisation")
        guide.geometry("620x520")
        guide.minsize(560, 420)
        guide.configure(bg="white")
        guide.transient(self)

        title = tk.Label(
            guide,
            text="Guide d'utilisation - Carnet de vol télépilote PRO",
            bg="white",
            fg="#073763",
            font=("Arial", 14, "bold"),
            anchor="w",
        )
        title.pack(fill=X, padx=16, pady=(14, 8))

        text_frame = tk.Frame(guide, bg="white")
        text_frame.pack(fill=BOTH, expand=True, padx=10, pady=(0, 10))
        text = tk.Text(text_frame, wrap="word", font=("Arial", 10), bg="white", relief="flat", padx=12, pady=8)
        scrollbar = ttk.Scrollbar(text_frame, orient=VERTICAL, command=text.yview)
        text.configure(yscrollcommand=scrollbar.set)
        text.pack(side=LEFT, fill=BOTH, expand=True)
        scrollbar.pack(side=RIGHT, fill=Y)
        text.insert(
            "1.0",
            "\n".join(
                [
                    "1. Ajouter un vol",
                    "Remplir les champs Date, Aéronef, Scénario, Domaine, Durée, Lieu et Observations, puis cliquer sur Ajouter / Modifier.",
                    "",
                    "2. Modifier une ligne",
                    "Cliquer sur une ligne du tableau, corriger les champs, puis cliquer sur Ajouter / Modifier.",
                    "",
                    "3. Nouveau",
                    "Vide les champs de saisie pour préparer une nouvelle ligne. Le tableau n'est pas supprimé.",
                    "",
                    "4. Supprimer",
                    "Supprime uniquement la ligne sélectionnée.",
                    "",
                    "5. Reset",
                    "Efface tout le carnet après confirmation.",
                    "",
                    "6. Sauvegarde",
                    "La sauvegarde est automatique : ajout, modification, suppression, reset et fermeture de l'application sont enregistrés sans bouton manuel.",
                    "Les données sont conservées dans le dossier utilisateur Windows de l'application. L'exe peut donc rester seul et indépendant.",
                    "",
                    "7. Exporter Excel",
                    "Crée un carnet Excel avec la répartition par scénario et par domaine.",
                    "",
                    "8. Lettre mission",
                    "Exporte une lettre de mission vierge intégrée dans l'application.",
                    "",
                    "9. CERFA",
                    "Exporte le CERFA PDF intégré dans l'application. Le bouton CERFA ne dépend pas du site internet : le document est dans l'exe.",
                    "",
                    "10. Onglets / liens internet du haut",
                    "Les boutons Alpha Tango, Cartes.gouv, SIA, METAR/TAF, Windy et Fly by sont des liens internet vers les différents sites utiles. Ils s'ouvrent dans le navigateur.",
                    "",
                    "11. Administrateur",
                    "Le mail de l'administrateur en bas de l'application est cliquable et ouvre une nouvelle rédaction avec le client mail par défaut.",
                ]
            ),
        )
        text.configure(state="disabled")

        ttk.Button(guide, text="Fermer", command=guide.destroy).pack(pady=(0, 12))

    def add_or_update(self) -> None:
        try:
            hours = parse_hours(self.vars["hours"].get())
        except ValueError:
            messagebox.showerror(APP_NAME, "Durée invalide. Exemple : 1, 1,5 ou 1:30.")
            return
        flight = Flight(
            date=self.vars["date"].get().strip(),
            aircraft=self.vars["aircraft"].get().strip(),
            scenario=normalize_scenario(self.vars["scenario"].get().strip()),
            domain=self.vars["domain"].get().strip() or "Autre",
            hours=hours,
            location=self.vars["location"].get().strip(),
            notes=self.vars["notes"].get().strip(),
        )
        if self.edit_index is None:
            self.flights.append(flight)
        else:
            self.flights[self.edit_index] = flight
        self.clear_form()
        self.refresh()
        self._write_data()

    def clear_form(self) -> None:
        self.edit_index = None
        self.vars["date"].set(date.today().strftime("%d/%m/%Y"))
        self.vars["aircraft"].set("")
        self.vars["scenario"].set("STS 01")
        self.vars["domain"].set("Photo")
        self.vars["hours"].set("1")
        self.vars["location"].set("")
        self.vars["notes"].set("")

    def selected_index(self) -> int | None:
        sel = self.tree.selection()
        if not sel:
            return None
        return int(sel[0])

    def load_selected(self, _event=None) -> None:
        idx = self.selected_index()
        if idx is None or idx >= len(self.flights):
            return
        self.edit_index = idx
        f = self.flights[idx]
        self.vars["date"].set(f.date)
        self.vars["aircraft"].set(f.aircraft)
        self.vars["scenario"].set(normalize_scenario(f.scenario))
        self.vars["domain"].set(f.domain)
        self.vars["hours"].set(str(f.hours).replace(".", ","))
        self.vars["location"].set(f.location)
        self.vars["notes"].set(f.notes)

    def delete_selected(self) -> None:
        idx = self.selected_index()
        if idx is None:
            return
        if messagebox.askyesno(APP_NAME, "Supprimer le vol sélectionné ?"):
            del self.flights[idx]
            self.clear_form()
            self.refresh()
            self._write_data()

    def reset_all(self) -> None:
        if messagebox.askyesno(APP_NAME, "Effacer toutes les lignes du carnet ?"):
            self.flights.clear()
            self.clear_form()
            self.refresh()
            self._write_data()

    def on_close(self) -> None:
        self._write_data()
        self.destroy()

    def refresh(self) -> None:
        self.tree.delete(*self.tree.get_children())
        for idx, f in enumerate(self.flights):
            scenario = normalize_scenario(f.scenario)
            self.tree.insert(
                "",
                END,
                iid=str(idx),
                values=(f.date, f.aircraft, scenario, f.domain, fmt_hours(f.hours), f.location, f.notes),
                tags=(scenario,),
            )

        total = sum(f.hours for f in self.flights)
        self.kpi_total.set(fmt_hours(total))
        self.kpi_count.set(str(len(self.flights)))
        domain_totals: dict[str, float] = {}
        scenario_totals: dict[str, float] = {s: 0.0 for s in SCENARIOS}
        for f in self.flights:
            domain_totals[f.domain] = domain_totals.get(f.domain, 0.0) + f.hours
            scenario = normalize_scenario(f.scenario)
            scenario_totals[scenario] = scenario_totals.get(scenario, 0.0) + f.hours

        self.summary_scenario.delete(*self.summary_scenario.get_children())
        for s in SCENARIOS:
            self.summary_scenario.insert("", END, values=(s, fmt_hours(scenario_totals.get(s, 0))))
        self.summary_domain.delete(*self.summary_domain.get_children())
        for domain, hours in sorted(domain_totals.items()):
            self.summary_domain.insert("", END, values=(domain, fmt_hours(hours)))

    def export_excel(self) -> None:
        default = f"Carnet_de_vol_telepilote_{YEAR}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        path = filedialog.asksaveasfilename(defaultextension=".xlsx", initialfile=default, filetypes=[("Excel", "*.xlsx")])
        if not path:
            return
        try:
            export_workbook(Path(path), self.flights, self.owner_name.get().strip())
            messagebox.showinfo(APP_NAME, "Export Excel terminé.")
        except Exception as exc:
            messagebox.showerror(APP_NAME, f"Export impossible : {exc}")

    def export_letter_template(self) -> None:
        source = resource_path(LETTER_TEMPLATE)
        if not source.exists():
            messagebox.showerror(APP_NAME, "Modèle de lettre de mission introuvable dans l'application.")
            return
        default = f"Lettre_de_mission_vierge_{datetime.now().strftime('%Y%m%d')}.docx"
        path = filedialog.asksaveasfilename(defaultextension=".docx", initialfile=default, filetypes=[("Word", "*.docx")])
        if not path:
            return
        try:
            shutil.copyfile(source, path)
            messagebox.showinfo(APP_NAME, "Lettre de mission vierge exportée.")
        except Exception as exc:
            messagebox.showerror(APP_NAME, f"Export impossible : {exc}")

    def export_cerfa_template(self) -> None:
        source = resource_path(CERFA_TEMPLATE)
        if not source.exists():
            messagebox.showerror(APP_NAME, "Document CERFA introuvable dans l'application.")
            return
        default = f"cerfa_15476-04_{datetime.now().strftime('%Y%m%d')}.pdf"
        path = filedialog.asksaveasfilename(defaultextension=".pdf", initialfile=default, filetypes=[("PDF", "*.pdf")])
        if not path:
            return
        try:
            shutil.copyfile(source, path)
            messagebox.showinfo(APP_NAME, "CERFA exporté.")
        except Exception as exc:
            messagebox.showerror(APP_NAME, f"Export impossible : {exc}")


    def export_mission_guide(self) -> None:
        default = f"Guide_preparation_mission_complet_{datetime.now().strftime('%Y%m%d')}.pdf"
        self.export_pdf_template(MISSION_GUIDE_TEMPLATE, default, "Guide de preparation mission")

    def export_consent_template(self) -> None:
        default = f"Attestation_consentement_vierge_{datetime.now().strftime('%Y%m%d')}.pdf"
        self.export_pdf_template(CONSENT_TEMPLATE, default, "Attestation de consentement")

    def export_pdf_template(self, template: str, default_name: str, label: str) -> None:
        source = resource_path(template)
        if not source.exists():
            messagebox.showerror(APP_NAME, f"{label} introuvable dans l'application.")
            return
        path = filedialog.asksaveasfilename(defaultextension=".pdf", initialfile=default_name, filetypes=[("PDF", "*.pdf")])
        if not path:
            return
        try:
            shutil.copyfile(source, path)
            messagebox.showinfo(APP_NAME, f"{label} exporte.")
        except Exception as exc:
            messagebox.showerror(APP_NAME, f"Export impossible : {exc}")


def export_workbook(path: Path, flights: list[Flight], owner_name: str = "") -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Répartition"
    build_summary_sheet(ws, flights, owner_name)
    for scenario in SCENARIOS:
        build_scenario_sheet(wb, scenario, [f for f in flights if normalize_scenario(f.scenario) == scenario])
    wb.save(path)


def base_styles():
    thin = Side(style="thin", color="808080")
    return {
        "border": Border(left=thin, right=thin, top=thin, bottom=thin),
        "title": Font(name="Arial", size=18, bold=True, color="1F4E78"),
        "header": Font(name="Arial", bold=True),
        "white": Font(name="Arial", bold=True, color="FFFFFF"),
        "blue": PatternFill("solid", fgColor="D9EAF7"),
        "dark": PatternFill("solid", fgColor="1F4E78"),
    }


def scenario_fill(scenario: str) -> PatternFill:
    return PatternFill("solid", fgColor=SCENARIO_COLORS.get(scenario, "#D9EAF7").replace("#", ""))


def scenario_font(scenario: str) -> Font:
    color = "FFFFFF" if scenario in {"A1", "A2", "A3"} else "000000"
    return Font(name="Arial", bold=True, color=color)


def build_summary_sheet(ws, flights: list[Flight], owner_name: str = "") -> None:
    st = base_styles()
    ws.sheet_view.showGridLines = False
    ws["A1"] = f"CARNET DE VOL TÉLÉPILOTE PRO DE : {owner_name or '________________'}"
    ws.merge_cells("A1:J1")
    ws["A1"].font = st["title"]
    ws["A1"].alignment = Alignment(horizontal="center")
    headers = ["Scénario", "Total"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(3, col, h)
        c.fill = st["dark"]
        c.font = st["white"]
        c.border = st["border"]
        c.alignment = Alignment(horizontal="center")
    scenario_totals = {s: 0.0 for s in SCENARIOS}
    domain_totals: dict[tuple[str, str], float] = {}
    domains = sorted({f.domain for f in flights} | set(DOMAINS))
    for f in flights:
        scenario = normalize_scenario(f.scenario)
        scenario_totals[scenario] = scenario_totals.get(scenario, 0.0) + f.hours
        domain_totals[(f.domain, scenario)] = domain_totals.get((f.domain, scenario), 0.0) + f.hours
    for row, s in enumerate(SCENARIOS, 4):
        ws.cell(row, 1, s)
        ws.cell(row, 2, scenario_totals.get(s, 0) / 24)
        ws.cell(row, 2).number_format = '[h]"h"mm'
        ws.cell(row, 1).fill = scenario_fill(s)
        ws.cell(row, 1).font = scenario_font(s)
        for col in [1, 2]:
            ws.cell(row, col).border = st["border"]
    ws.cell(12, 1, "TOTAL")
    ws.cell(12, 2, sum(scenario_totals.values()) / 24)
    ws.cell(12, 2).number_format = '[h]"h"mm'

    ws["D3"] = "Domaine d'activité"
    ws["D3"].fill = st["dark"]
    ws["D3"].font = st["white"]
    ws["D3"].border = st["border"]
    for idx, scenario in enumerate(SCENARIOS, 5):
        c = ws.cell(3, idx, scenario)
        c.fill = scenario_fill(scenario)
        c.font = scenario_font(scenario)
        c.border = st["border"]
    for row, domain in enumerate(domains, 4):
        ws.cell(row, 4, domain)
        ws.cell(row, 4).border = st["border"]
        for col, scenario in enumerate(SCENARIOS, 5):
            ws.cell(row, col, domain_totals.get((domain, scenario), 0) / 24)
            ws.cell(row, col).number_format = '[h]"h"mm'
            ws.cell(row, col).border = st["border"]
    for col, width in {"A": 18, "B": 14, "D": 28, "E": 14, "F": 14, "G": 14, "H": 14, "I": 14, "J": 14, "K": 14}.items():
        ws.column_dimensions[col].width = width


def build_scenario_sheet(wb: Workbook, scenario: str, flights: list[Flight]) -> None:
    st = base_styles()
    ws = wb.create_sheet(scenario)
    ws.sheet_view.showGridLines = False
    ws["A1"] = f"CARNET DE VOL {YEAR} - {scenario}"
    ws.merge_cells("A1:H1")
    ws["A1"].font = st["title"]
    ws["A1"].alignment = Alignment(horizontal="center")
    headers = ["N°", "Date", "Aéronef", "Domaine d'activité", "Durée (h)", "Cumul", "Lieu", "Observations"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(3, col, h)
        c.fill = st["dark"]
        c.font = st["white"]
        c.border = st["border"]
        c.alignment = Alignment(horizontal="center")
    cumul = 0.0
    for row, f in enumerate(flights, 4):
        cumul += f.hours
        values = [row - 3, f.date, f.aircraft, f.domain, f.hours, cumul / 24, f.location, f.notes]
        for col, value in enumerate(values, 1):
            c = ws.cell(row, col, value)
            c.border = st["border"]
            c.alignment = Alignment(vertical="center")
            if col == 6:
                c.number_format = '[h]"h"mm'
            if col == 5:
                c.number_format = "0.00"
    widths = [8, 14, 18, 26, 12, 12, 22, 34]
    for idx, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(idx)].width = width


def run_self_test() -> Path:
    sample = []
    for index, scenario in enumerate(SCENARIOS, 1):
        sample.append(
            Flight(
                date=f"{YEAR}-01-{index:02d}",
                aircraft="DJI Matrice",
                scenario=scenario,
                domain="Thermographie" if scenario == "STS 02" else "Photo",
                hours=float(index),
                location="Zone test",
                notes=f"Test export {scenario}",
            )
        )
    target = app_dir() / "CarnetVolTelepilotePRO_TEST.xlsx"
    export_workbook(target, sample, "TEST")
    letter_source = resource_path(LETTER_TEMPLATE)
    if not letter_source.exists():
        raise FileNotFoundError(f"Modèle introuvable : {letter_source}")
    shutil.copyfile(letter_source, app_dir() / "LettreMission_TEST.docx")
    cerfa_source = resource_path(CERFA_TEMPLATE)
    if not cerfa_source.exists():
        raise FileNotFoundError(f"CERFA introuvable : {cerfa_source}")
    shutil.copyfile(cerfa_source, app_dir() / "CERFA_TEST.pdf")
    mission_guide_source = resource_path(MISSION_GUIDE_TEMPLATE)
    if not mission_guide_source.exists():
        raise FileNotFoundError(f"Guide mission introuvable : {mission_guide_source}")
    shutil.copyfile(mission_guide_source, app_dir() / "GuideMission_TEST.pdf")
    consent_source = resource_path(CONSENT_TEMPLATE)
    if not consent_source.exists():
        raise FileNotFoundError(f"Attestation consentement introuvable : {consent_source}")
    shutil.copyfile(consent_source, app_dir() / "AttestationConsentement_TEST.pdf")
    return target


if __name__ == "__main__":
    if "--self-test" in sys.argv:
        path = run_self_test()
        print(f"SELF_TEST_OK {path}")
        sys.exit(0)
    CarnetApp().mainloop()
